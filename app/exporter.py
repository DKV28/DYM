"""Xuất kết quả trích xuất ra Excel (.xlsx) và JSON."""
import io
import json

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .schema import (
    NHAN_CA_NHAN,
    NHAN_CHUNG_CHI,
    NHAN_HOC_VAN,
    NHAN_LOAI_TAI_LIEU,
)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _ghi_tieu_de(ws, tieu_de: list[str]) -> None:
    ws.append(tieu_de)
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)


def _tu_dong_rong_cot(ws) -> None:
    for col in ws.columns:
        do_dai = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(do_dai + 2, 12), 45)


def xuat_json(ket_qua: list[dict]) -> bytes:
    """Trả về bytes JSON (UTF-8, giữ dấu tiếng Việt)."""
    return json.dumps(ket_qua, ensure_ascii=False, indent=2).encode("utf-8")


def xuat_excel(ket_qua: list[dict]) -> bytes:
    """Tạo workbook Excel với 3 sheet: Tổng hợp, Học vấn, Chứng chỉ.

    ket_qua: danh sách các bản ghi, mỗi bản ghi là dict:
        { "ten_file": str, "du_lieu": {<theo schema>}, "loi": str|None }
    """
    wb = Workbook()

    # ----- Sheet 1: Tổng hợp (mỗi tài liệu 1 dòng) -----
    ws = wb.active
    ws.title = "Tổng hợp"
    tieu_de = ["STT", "Tên file", "Loại tài liệu"]
    tieu_de += list(NHAN_CA_NHAN.values())
    tieu_de += ["Số bằng cấp", "Số chứng chỉ", "Ghi chú"]
    _ghi_tieu_de(ws, tieu_de)

    for i, ban_ghi in enumerate(ket_qua, start=1):
        d = ban_ghi.get("du_lieu") or {}
        ca_nhan = d.get("thong_tin_ca_nhan") or {}
        loai = d.get("loai_tai_lieu", "")
        dong = [
            i,
            ban_ghi.get("ten_file", ""),
            NHAN_LOAI_TAI_LIEU.get(loai, loai),
        ]
        dong += [ca_nhan.get(k, "") for k in NHAN_CA_NHAN]
        dong += [
            len(d.get("hoc_van") or []),
            len(d.get("chung_chi") or []),
            ban_ghi.get("loi") or d.get("ghi_chu", ""),
        ]
        ws.append(dong)
    _tu_dong_rong_cot(ws)
    ws.freeze_panes = "A2"

    # ----- Sheet 2: Học vấn (mỗi bằng 1 dòng) -----
    ws2 = wb.create_sheet("Học vấn")
    _ghi_tieu_de(ws2, ["Tên file", "Họ tên"] + list(NHAN_HOC_VAN.values()))
    for ban_ghi in ket_qua:
        d = ban_ghi.get("du_lieu") or {}
        ho_ten = (d.get("thong_tin_ca_nhan") or {}).get("ho_ten", "")
        for hv in d.get("hoc_van") or []:
            ws2.append(
                [ban_ghi.get("ten_file", ""), ho_ten]
                + [hv.get(k, "") for k in NHAN_HOC_VAN]
            )
    _tu_dong_rong_cot(ws2)
    ws2.freeze_panes = "A2"

    # ----- Sheet 3: Chứng chỉ (mỗi chứng chỉ 1 dòng) -----
    ws3 = wb.create_sheet("Chứng chỉ")
    _ghi_tieu_de(ws3, ["Tên file", "Họ tên"] + list(NHAN_CHUNG_CHI.values()))
    for ban_ghi in ket_qua:
        d = ban_ghi.get("du_lieu") or {}
        ho_ten = (d.get("thong_tin_ca_nhan") or {}).get("ho_ten", "")
        for cc in d.get("chung_chi") or []:
            ws3.append(
                [ban_ghi.get("ten_file", ""), ho_ten]
                + [cc.get(k, "") for k in NHAN_CHUNG_CHI]
            )
    _tu_dong_rong_cot(ws3)
    ws3.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
